from pathlib import Path
from collections import Counter, defaultdict
from typing import Dict, List
from datetime import datetime, time, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.core.parser import SessionRecord
from app.core.rules import Finding

ERROR_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')
WARNING_FILL = PatternFill(fill_type='solid', fgColor='FFF2CC')


def _calculate_evening_hours(start_dt: datetime, end_dt: datetime) -> float:
    if end_dt <= start_dt:
        return 0.0
    total_seconds = 0.0
    current_day = start_dt.date()
    last_day = end_dt.date()
    while current_day <= last_day:
        evening_start = datetime.combine(current_day, time(18, 0))
        evening_end = datetime.combine(current_day + timedelta(days=1), time(0, 0))
        overlap_start = max(start_dt, evening_start)
        overlap_end = min(end_dt, evening_end)
        if overlap_end > overlap_start:
            total_seconds += (overlap_end - overlap_start).total_seconds()
        current_day += timedelta(days=1)
    return total_seconds / 3600.0


def build_report(output_path: str | Path, sessions: List[SessionRecord], findings_by_session: Dict[str, List[Finding]]):
    details_rows = []
    summary_counter = Counter()
    full_rows = []
    evening_by_student = defaultdict(lambda: {'student_name': '', 'valid_session_count': 0, 'evening_hours': 0.0})

    for session in sessions:
        session_findings = findings_by_session.get(session.session_id, [])
        severities = {f.severity for f in session_findings}
        row_status = 'OK'
        if 'ERROR' in severities:
            row_status = 'ERROR'
        elif 'WARNING' in severities:
            row_status = 'WARNING'

        full_rows.append({
            'Excel Row': session.source_row_number,
            'STT': session.row_no,
            'Session ID': session.session_id,
            'Student ID': session.student_id,
            'Student Name': session.student_name_raw,
            'Teacher': session.teacher_name_raw,
            'Vehicle': session.vehicle_plate_raw,
            'Recognition': session.recognition,
            'Distance (km)': session.distance_km,
            'Duration (hours)': session.duration_hours,
            'Average Speed (km/h)': round(session.speed_kmh, 2),
            'Start Time': session.start_time.strftime('%d/%m/%Y %H:%M'),
            'End Time': session.end_time.strftime('%d/%m/%Y %H:%M'),
            'Status': row_status,
            'Issues': '; '.join(f.message for f in session_findings),
        })

        if row_status != 'ERROR':
            evening_hours = _calculate_evening_hours(session.start_time, session.end_time)
            if evening_hours > 0:
                student_entry = evening_by_student[session.student_id]
                student_entry['student_name'] = session.student_name_raw
                student_entry['valid_session_count'] += 1
                student_entry['evening_hours'] += evening_hours

        for finding in session_findings:
            summary_counter[finding.severity] += 1
            details_rows.append({
                'Excel Row': session.source_row_number,
                'STT': session.row_no,
                'Session ID': session.session_id,
                'Student ID': session.student_id,
                'Student Name': session.student_name_raw,
                'Rule': finding.rule,
                'Severity': finding.severity,
                'Message': finding.message,
                'Value': finding.value or '',
            })

    summary_df = pd.DataFrame([
        {'Metric': 'Total Sessions', 'Value': len(sessions)},
        {'Metric': 'Error Count', 'Value': summary_counter['ERROR']},
        {'Metric': 'Warning Count', 'Value': summary_counter['WARNING']},
        {'Metric': 'Sessions With Findings', 'Value': sum(1 for s in sessions if findings_by_session.get(s.session_id))},
    ])
    details_df = pd.DataFrame(details_rows)
    full_df = pd.DataFrame(full_rows)
    evening_df = pd.DataFrame([
        {
            'Student ID': student_id,
            'Student Name': values['student_name'],
            'Valid Evening Sessions Count': values['valid_session_count'],
            'Evening Riding Time (hours)': round(values['evening_hours'], 2),
        }
        for student_id, values in sorted(evening_by_student.items())
    ])
    if evening_df.empty:
        evening_df = pd.DataFrame(columns=['Student ID', 'Student Name', 'Valid Evening Sessions Count', 'Evening Riding Time (hours)'])

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        details_df.to_excel(writer, sheet_name='Findings', index=False)
        full_df.to_excel(writer, sheet_name='Full Data', index=False)
        evening_df.to_excel(writer, sheet_name='Evening Summary', index=False)

    _highlight_workbook(output_path)


def _highlight_workbook(path: str | Path):
    wb = load_workbook(path)

    findings_ws = wb['Findings']
    severity_col = None
    for idx, cell in enumerate(findings_ws[1], start=1):
        if cell.value == 'Severity':
            severity_col = idx
            break
    if severity_col:
        for row in findings_ws.iter_rows(min_row=2):
            severity = row[severity_col - 1].value
            fill = ERROR_FILL if severity == 'ERROR' else WARNING_FILL if severity == 'WARNING' else None
            if fill:
                for cell in row:
                    cell.fill = fill

    full_ws = wb['Full Data']
    status_col = None
    for idx, cell in enumerate(full_ws[1], start=1):
        if cell.value == 'Status':
            status_col = idx
            break
    if status_col:
        for row in full_ws.iter_rows(min_row=2):
            status = row[status_col - 1].value
            fill = ERROR_FILL if status == 'ERROR' else WARNING_FILL if status == 'WARNING' else None
            if fill:
                for cell in row:
                    cell.fill = fill

    wb.save(path)
